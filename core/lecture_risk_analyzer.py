"""行课数据风险分析器 — 沟通优先级排序 + 回访建议话术"""
import os
import sys
import re
import pandas as pd
from openai import OpenAI
from concurrent.futures import ThreadPoolExecutor, as_completed

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import OPENAI_API_KEY, OPENAI_BASE_URL, MODEL_NAME

client = OpenAI(api_key=OPENAI_API_KEY, base_url=OPENAI_BASE_URL)
MAX_WORKERS = 20


def get_client() -> OpenAI:
    """统一使用 config.py 中的硬编码配置，确保所有模块用同一个 Key"""
    return OpenAI(api_key=OPENAI_API_KEY, base_url=OPENAI_BASE_URL)


from core.prompt_utils import append_constraints


def load_prompt(template_name: str) -> str:
    prompt_dir = os.path.join(os.path.dirname(__file__), "prompts")
    filepath = os.path.join(prompt_dir, template_name)
    with open(filepath, "r", encoding="utf-8") as f:
        return append_constraints(f.read())


def _parse_val(val) -> float:
    """从单元格提取数值"""
    if pd.isna(val):
        return None
    s = str(val).strip()
    if not s or s in ["无", "-", "N/A", "n/a", "未提交", "未完成", "未布置"]:
        return None
    try:
        # 尝试提取百分比或数字
        m = re.search(r'(\d+(?:\.\d+)?)', s)
        if m:
            return float(m.group(1))
    except:
        pass
    return None


def _is_effective(row_data, lec, attendance_rates):
    """判断是否有效听课
    Returns:
        True  - 明确有效听课
        False - 明确无效听课
        None  - 无法判断（数据缺失或未到课）
    """
    effective_col = lec["cols"].get("是否有效听课")
    if effective_col and effective_col in row_data:
        val = str(row_data[effective_col]).strip()
        if val in ["是", "1", "True", "true", "有效"]:
            return True
        if val in ["否", "0", "False", "false", "无效"]:
            return False
    return None  # 没有"是否有效听课"列，无法判断


def analyze_student_risk(row_data: dict, lectures: list, attendance_rates: dict) -> dict:
    """
    分析单个学员的5维度风险指标
    返回沟通优先级分数和详细数据
    """
    metrics = {
        "有效听课": {"count": 0, "total": 0, "rate": 0},
        "听课时长": {"values": [], "avg": 0},
        "答题正确率": {"values": [], "avg": 0},
        "练习提交": {"count": 0, "total": 0, "rate": 0},
        "练习得分": {"values": [], "avg": 0},
    }
    # 4类异常明细（用于导出Excel）
    anomalies = {
        "无效听课": [],       # [(讲次, 标题)]
        "答题低于70": [],     # [(讲次, 正确率)]
        "未提交练习": [],     # [(讲次,)]
        "练习低于70": [],     # [(讲次, 得分)]
    }
    bad_lectures = []
    good_lectures = []

    for lec in lectures:
        lec_name = lec["lecture"]
        lec_title = lec["title"]
        cols = lec["cols"]

        # 不跳过任何讲次，所有讲次都参与统计
        metrics["有效听课"]["total"] += 1

        # 1. 有效听课
        is_eff = _is_effective(row_data, lec, attendance_rates)
        if is_eff is True:
            metrics["有效听课"]["count"] += 1
        elif is_eff is False:
            # 只有明确无效才计入异常
            anomalies["无效听课"].append((lec_name, lec_title))
            bad_lectures.append({
                "lecture": lec_name, "title": lec_title,
                "reason": "无效听课", "details": []
            })
        # is_eff is None 表示未开课，不计入任何统计

        # 2. 听课时长
        dur_col = cols.get("听课时长")
        if dur_col and dur_col in row_data:
            dur = _parse_val(row_data[dur_col])
            if dur is not None and dur > 0:
                metrics["听课时长"]["values"].append(dur)

        # 3. 答题正确率（阈值70%）
        acc_col = cols.get("直播答题正确率")
        if acc_col and acc_col in row_data:
            acc = _parse_val(row_data[acc_col])
            if acc is not None and acc > 0:
                metrics["答题正确率"]["values"].append(acc)
                if acc < 70:
                    anomalies["答题低于70"].append((lec_name, acc))

        # 4. 练习提交
        practice_col = cols.get("练习状态")
        if practice_col and practice_col in row_data:
            pv = str(row_data[practice_col]).strip()
            if pv not in ["未提交", "未完成", "0", "", "nan", "None"]:
                metrics["练习提交"]["count"] += 1
            else:
                anomalies["未提交练习"].append((lec_name,))
            metrics["练习提交"]["total"] += 1

        # 5. 练习得分（阈值70分）
        score_col = cols.get("练习得分")
        if score_col and score_col in row_data:
            sc = _parse_val(row_data[score_col])
            if sc is not None and sc > 0:
                metrics["练习得分"]["values"].append(sc)
                if sc < 70:
                    anomalies["练习低于70"].append((lec_name, sc))

    # 计算各项比率
    if metrics["有效听课"]["total"] > 0:
        metrics["有效听课"]["rate"] = metrics["有效听课"]["count"] / metrics["有效听课"]["total"]
    if metrics["听课时长"]["values"]:
        metrics["听课时长"]["avg"] = sum(metrics["听课时长"]["values"]) / len(metrics["听课时长"]["values"])
    if metrics["答题正确率"]["values"]:
        metrics["答题正确率"]["avg"] = sum(metrics["答题正确率"]["values"]) / len(metrics["答题正确率"]["values"])
    if metrics["练习提交"]["total"] > 0:
        metrics["练习提交"]["rate"] = metrics["练习提交"]["count"] / metrics["练习提交"]["total"]
    if metrics["练习得分"]["values"]:
        metrics["练习得分"]["avg"] = sum(metrics["练习得分"]["values"]) / len(metrics["练习得分"]["values"])

    # 综合风险分 — 优先级规则
    # P1: 有无效听课 或 未提交练习（硬性红线）
    # P2-P4: 按答题正确率 + 练习分数排序
    invalid_count = len(anomalies.get("无效听课", []))
    unpractice_count = len(anomalies.get("未提交练习", []))
    low_acc = anomalies.get("答题低于70", [])
    low_score = anomalies.get("练习低于70", [])

    if invalid_count > 0 or unpractice_count > 0:
        # 直接P1，按严重程度细分风险分
        risk_score = 70 + min(invalid_count + unpractice_count, 6) * 5
        priority = "P1-紧急"
    elif low_acc or low_score:
        # P2-P3: 按正确率和分数加权
        acc_score = 0
        if low_acc:
            acc_avg = sum(x[1] for x in low_acc) / len(low_acc)
            acc_score = (70 - acc_avg) / 70 * 20 + len(low_acc) * 5
        score_score = 0
        if low_score:
            sc_avg = sum(x[1] for x in low_score) / len(low_score)
            score_score = (70 - sc_avg) / 70 * 15 + len(low_score) * 4
        risk_score = acc_score + score_score
        # 有效听课率偏低额外加分
        if metrics["有效听课"]["total"] > 0 and metrics["有效听课"]["rate"] < 0.8:
            risk_score += (1 - metrics["有效听课"]["rate"]) * 15
        # 练习提交率偏低额外加分
        if metrics["练习提交"]["total"] > 0 and metrics["练习提交"]["rate"] < 0.8:
            risk_score += (1 - metrics["练习提交"]["rate"]) * 10
        risk_score = min(round(risk_score, 1), 69)
        priority = "P2-高" if risk_score >= 35 else "P3-中"
    else:
        risk_score = 0
        priority = "P4-低"

    return {
        "risk_score": round(risk_score, 1),
        "metrics": metrics,
        "bad_lectures": bad_lectures,
        "anomalies": anomalies,
        "priority": priority,
    }


def build_risk_context(risk_result: dict, student_name: str) -> str:
    """将风险分析结果组织为LLM可读的文本"""
    m = risk_result["metrics"]
    lines = [f"【学员】{student_name}"]
    lines.append(f"【综合风险分】{risk_result['risk_score']}/100（优先级：{risk_result['priority']}）")
    lines.append("")

    lines.append("【5维度指标】")
    if m["有效听课"]["total"] > 0:
        lines.append(f"  有效听课率：{m['有效听课']['count']}/{m['有效听课']['total']} = {m['有效听课']['rate']*100:.0f}%")
    else:
        lines.append("  有效听课率：无数据")
    if m["听课时长"]["values"]:
        lines.append(f"  平均听课时长：{m['听课时长']['avg']:.0f}分钟")
    if m["答题正确率"]["values"]:
        lines.append(f"  平均答题正确率：{m['答题正确率']['avg']:.0f}%")
    if m["练习提交"]["total"] > 0:
        lines.append(f"  练习提交率：{m['练习提交']['count']}/{m['练习提交']['total']} = {m['练习提交']['rate']*100:.0f}%")
    if m["练习得分"]["values"]:
        lines.append(f"  平均练习得分：{m['练习得分']['avg']:.0f}分")

    if risk_result["bad_lectures"]:
        lines.append("")
        lines.append("【需关注的问题讲次】")
        for bl in risk_result["bad_lectures"]:
            detail_str = "，".join(bl["details"]) if bl["details"] else ""
            lines.append(f"  {bl['lecture']}《{bl['title']}》：{bl['reason']}{'（' + detail_str + '）' if detail_str else ''}")

    return "\n".join(lines)


def generate_risk_script(risk_result: dict, student_name: str) -> str:
    """调用LLM生成沟通建议+回访话术"""
    template = load_prompt("lecture_risk_script.txt")
    context = build_risk_context(risk_result, student_name)
    prompt = template.format(student_name=student_name, risk_context=context)

    try:
        response = client.chat.completions.create(
            model=MODEL_NAME,
            temperature=0.3,
            max_tokens=1800,
            messages=[
                {"role": "system", "content": "你是高途二讲老师回访沟通助手。根据学员数据给出具体沟通建议和话术，禁止编造。输出简洁精炼，不要过于冗长。"},
                {"role": "user", "content": prompt}
            ]
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        return f"生成失败：{e}"


def batch_analyze_risk(df: pd.DataFrame, lectures: list, attendance_rates: dict,
                       name_col: str, id_col: str = None, progress_callback=None) -> list:
    """
    批量分析所有学员风险，按优先级排序，返回结果列表
    """
    rows_data = df.to_dict(orient="records")
    total = len(rows_data)
    results = []

    for idx, row_data in enumerate(rows_data):
        name = str(row_data.get(name_col, "")) if row_data.get(name_col) and pd.notna(row_data[name_col]) else "未知"
        sid = str(row_data.get(id_col, "")) if id_col and row_data.get(id_col) and pd.notna(row_data[id_col]) else ""

        risk = analyze_student_risk(row_data, lectures, attendance_rates)

        m = risk["metrics"]
        results.append({
            "排名": 0,
            "学员姓名": name,
            "学员ID": sid,
            "风险分": risk["risk_score"],
            "优先级": risk["priority"],
            "有效听课率": f"{m['有效听课']['rate']*100:.0f}%" if m['有效听课']['total'] > 0 else "无数据",
            "平均正确率": f"{m['答题正确率']['avg']:.0f}%" if len(m['答题正确率']['values']) > 0 else "无数据",
            "练习提交率": f"{m['练习提交']['rate']*100:.0f}%" if m['练习提交']['total'] > 0 else "无数据",
            "平均得分": f"{m['练习得分']['avg']:.0f}" if len(m['练习得分']['values']) > 0 else "无数据",
            "问题讲次": sum(len(v) for v in risk["anomalies"].values()),
            "_risk_detail": risk,
            "_row_data": row_data,
        })
        if progress_callback:
            progress_callback(idx + 1, total)

    results.sort(key=lambda x: x["风险分"], reverse=True)
    for i, r in enumerate(results):
        r["排名"] = i + 1
    return results


def batch_generate_scripts(results: list, progress_callback=None) -> list:
    """为排序后的学员并发生成沟通话术"""
    total = len(results)

    def _task(idx, item):
        try:
            script = generate_risk_script(item["_risk_detail"], item["学员姓名"])
            return idx, script
        except Exception as e:
            return idx, f"生成失败：{e}"

    scripts = [""] * total
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(_task, i, r): i for i, r in enumerate(results)}
        for future in as_completed(futures):
            idx, script = future.result()
            scripts[idx] = script
            if progress_callback:
                progress_callback(idx + 1, total)

    return scripts


def export_risk_excel(results: list, lectures: list, attendance_rates: dict,
                       output_path: str, df: pd.DataFrame = None) -> str:
    """
    导出风险分析Excel，直接openpyxl写入（速度优化）
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    def _lecture_sort_key(lec_name):
        m = re.search(r'(\d+)', lec_name)
        return int(m.group(1)) if m else 0

    def _fmt_lecture_list(anomaly_list, show_value=True):
        if not anomaly_list:
            return "-"
        sorted_list = sorted(anomaly_list, key=lambda x: _lecture_sort_key(x[0]))
        count = len(sorted_list)
        if show_value and len(sorted_list[0]) > 1 and sorted_list[0][1] is not None:
            values = [x[1] for x in sorted_list]
            avg = sum(values) / len(values)
            detail = "、".join(f"{n} {v:.0f}%" for n, v in sorted_list)
            return f"{count}讲（均{avg:.0f}%）：{detail}"
        return f"{count}讲：{'、'.join(x[0] for x in sorted_list)}"

    def _fmt_score_list(anomaly_list):
        if not anomaly_list:
            return "-"
        sorted_list = sorted(anomaly_list, key=lambda x: _lecture_sort_key(x[0]))
        count = len(sorted_list)
        values = [x[1] for x in sorted_list]
        avg = sum(values) / len(values)
        detail = "、".join(f"{n} {v:.0f}分" for n, v in sorted_list)
        return f"{count}讲（均{avg:.0f}分）：{detail}"

    def _gen_script(item):
        risk = item["_risk_detail"]
        anomalies = risk["anomalies"]
        invalid_lec = anomalies.get("无效听课", [])
        low_acc_lec = anomalies.get("答题低于70", [])
        unpractice_lec = anomalies.get("未提交练习", [])
        low_score_lec = anomalies.get("练习低于70", [])
        tasks = []
        if invalid_lec:
            lec_names = "、".join(sorted([x[0] for x in invalid_lec], key=_lecture_sort_key))
            tasks.append(f"补看{lec_names}回放")
        if unpractice_lec:
            lec_names = "、".join(sorted([x[0] for x in unpractice_lec], key=_lecture_sort_key))
            tasks.append(f"完成{lec_names}未提交练习")
        if low_score_lec:
            lec_names = "、".join(sorted([x[0] for x in low_score_lec], key=_lecture_sort_key))
            tasks.append(f"重做{lec_names}低分练习并订正")
        if not tasks:
            return "暂无异常，继续保持。"
        ruku = "【入手】" + "，".join(tasks) + "。"
        parts = []
        if invalid_lec:
            parts.append(f"{len(invalid_lec)}讲听课未达标")
        if low_acc_lec:
            avg_acc = sum(x[1] for x in low_acc_lec) / len(low_acc_lec)
            parts.append(f"{len(low_acc_lec)}讲答题偏低（均{avg_acc:.0f}%）")
        if unpractice_lec:
            parts.append(f"{len(unpractice_lec)}讲练习未交")
        if low_score_lec:
            avg_sc = sum(x[1] for x in low_score_lec) / len(low_score_lec)
            parts.append(f"{len(low_score_lec)}讲练习低分（均{avg_sc:.0f}分）")
        goutong = "【沟通】孩子目前" + "、".join(parts) + "，请家长配合督促。"
        zhongyao = "【重要性】问题较多需尽快补上，否则影响后续课程吸收。" if len(tasks) >= 3 else "【重要性】薄弱环节及时补上，否则影响后续知识吸收。"
        wancheng = "【完成时间】建议2周内完成，完成后反馈老师。"
        return "\n".join([ruku, goutong, zhongyao, wancheng])

    # 预计算所有行数据
    headers = ["排名", "姓名", "优先级",
               "无效听课\n（讲次）", "答题<70%\n（讲次·正确率）",
               "未提交练习\n（讲次）", "练习<70分\n（讲次·得分）", "参考话术"]
    rows = []
    for item in results:
        anomalies = item["_risk_detail"]["anomalies"]
        rows.append([
            item["排名"],
            item["学员姓名"],
            item["优先级"],
            _fmt_lecture_list(anomalies.get("无效听课", []), show_value=False),
            _fmt_lecture_list(anomalies.get("答题低于70", []), show_value=True),
            _fmt_lecture_list(anomalies.get("未提交练习", []), show_value=False),
            _fmt_score_list(anomalies.get("练习低于70", [])),
            _gen_script(item),
        ])

    # 直接openpyxl写入
    wb = Workbook()
    ws = wb.active
    ws.title = "沟通优先级"

    # 颜色定义
    p1_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    p2_fill = PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid")
    p3_fill = PatternFill(start_color="FFD43B", end_color="FFD43B", fill_type="solid")
    p4_fill = PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid")
    header_fill = PatternFill(start_color="2D3436", end_color="2D3436", fill_type="solid")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 写表头
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 写数据行
    for row_idx, row_data in enumerate(rows, 2):
        priority_val = str(row_data[2])  # 优先级在索引2
        is_p1 = "P1" in priority_val
        is_p2 = "P2" in priority_val
        row_fill = p1_fill if is_p1 else p2_fill if is_p2 else p3_fill if "P3" in priority_val else p4_fill

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            if col_idx == 3:
                cell.fill = row_fill
                cell.font = Font(name="微软雅黑", bold=True, size=10,
                                 color="FFFFFF" if is_p1 or is_p2 else "333333")
                cell.alignment = center_align
            elif col_idx == 1:
                cell.alignment = center_align
            elif col_idx == 8:
                cell.alignment = wrap_align
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 交替底色
        if row_idx % 2 == 0:
            alt_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            for col_idx in [1, 2, 4, 5, 6, 7, 8]:
                if col_idx != 3:
                    ws.cell(row=row_idx, column=col_idx).fill = alt_fill

        ws.row_dimensions[row_idx].height = 60

    # 列宽
    col_widths = {"A": 6, "B": 10, "C": 10, "D": 22, "E": 55, "F": 22, "G": 40, "H": 55}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    wb.save(output_path)
    return output_path