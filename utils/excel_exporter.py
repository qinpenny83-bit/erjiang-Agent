"""Excel导出器 — 将分析结果和报告导出为Excel文件"""
import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def export_analysis_to_excel(analysis_result: dict, output_path: str):
    """导出学情分析结果为Excel（支持动态字段）"""
    # 提取所有可能的列
    all_keys = set()
    for student in analysis_result["students"]:
        all_keys.update(student.keys())

    # 排除不需要导出的列（内部字段和复杂类型）
    exclude_keys = {"风险详情", "raw_data", "classification", "_profile", "评分依据", "风险触发", "缺失维度", "家长信号", "recognition", "_profile"}
    export_keys = [k for k in all_keys if k not in exclude_keys]

    # 构建DataFrame，处理复杂类型
    rows = []
    for student in analysis_result["students"]:
        row = {}
        for k in export_keys:
            val = student.get(k, "")
            # 将列表/字典转为可读字符串
            if isinstance(val, list):
                if val and isinstance(val[0], dict):
                    # 各维度风险 → 浓缩为文本
                    parts = []
                    for d in val:
                        dim_name = d.get("维度", "")
                        dim_level = d.get("风险等级", "")
                        dim_evidence = d.get("判断依据", "")
                        parts.append(f"{dim_name}({dim_level}):{dim_evidence}")
                    row[k] = " | ".join(parts)
                else:
                    # 普通列表 → 分号连接
                    row[k] = "；".join(str(v) for v in val)
            elif isinstance(val, dict):
                row[k] = str(val)
            else:
                row[k] = val
        rows.append(row)

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="学生分层列表", index=False)

        # 写入分层统计
        stats_df = pd.DataFrame([
            {"分层": tier, "人数": count}
            for tier, count in analysis_result["tier_stats"].items()
        ])
        stats_df.to_excel(writer, sheet_name="分层统计", index=False)

        # 写入预警数据（如果有）
        if "warnings" in analysis_result:
            warnings_df = pd.DataFrame([
                {"预警类型": k, "人数": v}
                for k, v in analysis_result["warnings"].items()
            ])
            warnings_df.to_excel(writer, sheet_name="预警统计", index=False)

    return output_path


def export_reports_to_excel(reports: list[dict], output_path: str):
    """导出批量报告为Excel（每个学生一个sheet）"""
    wb = Workbook()
    wb.remove(wb.active)

    # 定义样式
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    tier_fills = {
        "S": PatternFill(start_color="FF4B4B", end_color="FF4B4B", fill_type="solid"),
        "A": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
        "B": PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
        "C": PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid"),
    }
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for report in reports:
        name = report.get("学生姓名", report.get("name", "未知"))
        tier = report.get("分层", report.get("分类", "B"))
        # sheet名称限制31字符
        sheet_name = name[:31] if len(name) > 31 else name
        ws = wb.create_sheet(title=sheet_name)

        # 标题行
        ws.merge_cells("A1:B1")
        ws["A1"] = f"{name} - 学情反馈报告"
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 35

        # 分层标签
        ws["A3"] = "分层等级"
        ws["A3"].font = Font(bold=True)
        ws["B3"] = tier
        ws["B3"].fill = tier_fills.get(tier, PatternFill())
        ws["B3"].font = Font(bold=True, color="FFFFFF")

        # 报告内容
        ws["A5"] = "学情报告"
        ws["A5"].font = Font(bold=True)
        ws.merge_cells("A6:B30")
        report_text = report.get("正式报告", report.get("完整输出", report.get("学情报告", report.get("report", ""))))
        ws["A6"] = report_text
        ws["A6"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[6].height = 300

        # 设置列宽
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 80

    wb.save(output_path)
    return output_path
